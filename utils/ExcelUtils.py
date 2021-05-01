import openpyxl


def getRowCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_row


def getColCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_column


def readData(file, sheetName, rowno, colno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(rowno, colno).value


def writeData(file, sheetName, rowno, colno, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(rowno, colno).value = data
    workbook.save(file)


def writelistoflist(file, sheetName, data):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = sheetName
    row = 1
    for l in data:
        col = 1
        for d in l:
            if d:
                sheet.cell(row, col).value = d
            col += 1
        row += 1
    wb.save(file)


def singlerow(file, sheetName, rownum, colnum):
    single_row = []
    for i in range(1, colnum + 1):
        data = readData(file, sheetName, rownum, i)
        single_row.append(data)
    return single_row
